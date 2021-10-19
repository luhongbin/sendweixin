# coding=utf-8
from __future__ import unicode_literals
import pymssql  as pymssql
import time,os,datetime
import xlrd
import psycopg2.extras
import hashlib
import cx_Oracle
import openpyxl
import xlwt
import traceback
import smtplib
from email.mime.text import MIMEText
from email.utils import formataddr
import pymssql as pymssql
import requests,urllib
from PIL import Image
from ctypes import *
import win32gui
import win32clipboard as w
import win32api
import win32con
import json,xlrd,os
import datetime,time,re
from wxpy import *
from bs4 import BeautifulSoup
import math
import win32com.client
import pythoncom

os.environ['NLS_LANG'] = 'SIMPLIFIED CHINESE_CHINA.UTF8'

DBNAME = 'active'
DBHOST = '???.247.90.???'
DBUSER = 'xxx'
DBPASS = 'xxx@xxx'
DBPORT = 25252
# bot = Bot()

def get_t100():
    DBNAME1 = 'CIMS'
    DBHOST1 = 'localhost'
    DBUSER1 = 'openpg'
    DBPASS1 = 'xxxxxx'
    DBPORT1 = 5432
    my_sender = 'system@umenb.com'  # 发件人邮箱账号
    my_pass = 'seqhsqumamnxcafa'  # 发件人邮箱密码(当时申请smtp给的口令)

    conn1 = psycopg2.connect("host=%s port=%s dbname=%s user=%s password=%s" % (DBHOST1, DBPORT1, DBNAME1, DBUSER1, DBPASS1))
    openpg = conn1.cursor()
    conn2 = cx_Oracle.connect('dsdata/dsdata@192.168.0.5:1521/TOPPRD')
    cur = conn2.cursor()

    connhr = pymssql.connect(host='192.168.0.23', user='xxx', password='xxx', database='HRMDB',charset='utf8')
    curhr = connhr.cursor()
    ##########从HR中获取代码设置为POSTGRESQL供应商平台账户,密码yxh123,完成授权并发送邮件
    xxx = "select yggh 工号,ygxm 姓名,Email,bmmc 部门,zwmc 职位,Code 直属主管工号,CnName 直属主管姓名 from cwygxx order by yggh"
    curhr.execute(xxx)
    res = curhr.fetchall()
    for tao in res:
        r0 = tao[0]
        r1 = tao[1]
        r2 = tao[2]

        sql = "select username from  auth_user  where  username=('%s')" % (r0)
        openpg.execute(sql)
        conn1.commit()  # 执行插入
        res = openpg.fetchone()
        if not res:
            print(r0, r1, r2)
            sql = "insert into auth_user (password,  is_superuser, username, first_name,  email, is_staff, is_active,last_name,date_joined) values ('pbkdf2_sha256$120000$CqYFUczpD3Pq$PUiF/M9rJ9TzEW7OIXCsXPBYS2Xs0V1Dr8Bs+thKgf8=',false ,'%s','%s','%s',true ,true,'',LOCALTIMESTAMP )" % (
            r0, r1, r2)
            openpg.execute(sql)
            conn1.commit()  # 执行插入
        r3 = tao[3]
        r4 = tao[4]
        r5 = tao[5]
        r6 = tao[6]
        sql = "update auth_user set bumen=('%s'), zhiwei=('%s'), leadercode=('%s'), leadername=('%s'),email=('%s') WHERE username=('%s')" % (r3, r4,r5, r6,r0,r2)
        openpg.execute(sql)
        conn1.commit()  # 执行插
    connhr.close


    ##########从T100询价单中获取数据转让POSTGRESQL
    xxx = "select pmdgsite 公司别,pmdgdocno 询价单号,pmdgseq 项次,decode(pmdg001,'Y','外加工','采购') 属性,pmdg002 供应商编号,pmaal004 供应商简称,pmdg003 品号,imaal003 品名,imaal004 规格,pmdgua006 报价分类, pmdg007 询价数量,pmdg008 询价单位,pmdg009 分量计价否,pmdg011 税率,pmdgud014 最小包装量,pmdg013 最低采购量,pmdgua007 产品报价含税,pmdgud013 理论重量首次报价,pmdgud012 模具费报价含税,pmdg017 有效日期,pmdg030 计价公式及备注,pmdf002 工号,ooag011 姓名,pmdgud017,pmdgud005,pmdgud018,pmdgud019,pmdgud020,imaa016 毛重,imaaua113 抗UV ,imaaua114 阻燃等级 ,imaaua142 材质要求  from dsdata.pmdg_t left join dsdata.pmdf_t on pmdfent=pmdgent and pmdfdocno=pmdgdocno left join dsdata.pmaal_t on pmaalent=pmdgent and pmaal001=pmdg002 and pmaal002='zh_CN' left join dsdata.imaal_t on imaal001 =pmdg003 and imaalent=pmdgent  and imaal002='zh_CN' left join dsdata.ooag_t  on pmdf002=ooag001 and ooagent=pmdgent  left join dsdata.imaa_t on imaa001 =pmdg003 and imaaent=pmdgent  where pmdgent ='60' and  pmdfstus='N' and pmdg003 is not null and pmdg002 is not null and pmdfud002='Y'"
    cur.execute(xxx)
    res = cur.fetchall()
    for tao in res:
        r0 = tao[0]
        if r0=='Y1':
            r0='Y1-耀泰'
        if r0=='Y3':
            r0='Y3-颐道'
        r1 = tao[1]
        r2 = str(tao[2])
        print(r0, r1, r2)
        r4 = tao[4]
        r3 = tao[3]
        r5 = tao[5]
        r6 = tao[6]
        r7 = tao[7]
        if r7 is None:
            r7 = ''
        r8 = tao[8]
        if r8 is None:
            r8 = ''
        r9 = tao[9]
        if r9 is None:
            r9 = 1
        if r9 == '0':
            r9 = 1
        else:
            r9 = 2

        sql = "select pmdgsite from  basedata_pmdg  where  pmdgsite=('%s') and pmdgdocno=('%s') and  pmdgseq=('%s')  and  pmdg002=('%s')" % (r0,r1,r2,r4)
        openpg.execute(sql)
        #conn1.commit()  # 执行插入
        res = openpg.fetchone()
        if not res:

            sql = "insert into basedata_pmdg (pmdgsite,  pmdgdocno, pmdgseq,pmdg002,status) values ('%s','%s','%s','%s','0')" % (r0, r1, r2, r4)
            openpg.execute(sql)
            conn1.commit()  # 执行插入
            # try:
            print("邮件开始")
            sql = "select id from  basedata_pmdg  where  pmdgsite=('%s') and pmdgdocno=('%s') and  pmdgseq=('%s') and  pmdg002=('%s')" % (
            r0, r1, r2, r4)
            openpg.execute(sql)
            # conn1.commit()  # 执行插入
            row = openpg.fetchone()
            if row == None:
                product_id = str(row[0])
            else:
                product_id = str(row[0])
            from email.mime.text import MIMEText
            from email.utils import formataddr
            import smtplib
            sql = "select email from  auth_user  where  username=('%s')" % (r4)
            openpg.execute(sql)
            res = openpg.fetchall()
            for tao1 in res:
                ttx = tao1[0]
                if ttx == None:
                    ttx = ''
                print(ttx)
            my_user = 'finance@lutec.net'  # 收件人邮箱账号，我这边发送给自己
            my_sender = '123660733@qq.com'  # 发件人邮箱账号
            my_pass = 'xxxx'  # 发件人邮箱密码(当时申请smtp给的口令)
            try:
                msg = MIMEText(
                    '亲爱的供应商【' + r5 + '】你好：<br><p>我们邀请您使用供应商管理平台：</p><p><a href="https://www.????.com/admin/basedata/pmdg/' + product_id + '">????(集团)管理平台</a><br>对品名：' + r7 + '，规格：' + r8 + '<br>及时报价。<br><hr>技术支持：宁波耀泰(集团)，<a href=mailto:finance@lutec.net>客服邮箱</a></p>',
                    'html', 'utf-8')
            except Exception:
                print('亲爱的供应商【' + r5 + '】你好：<br><p>我们邀请您使用供应商管理平台：</p><p><a href="https://www.????.com/admin/basedata/pmdg/' + product_id + '">????(集团)管理平台</a><br>对品名：' +r7 + '，规格：' + r8 + '<br>及时报价。<br><hr>技术支持：宁波耀泰(集团)，<a href=mailto:finance@lutec.net>客服邮箱</a></p>')
            msg['From'] = formataddr(["发件人昵称", my_sender])  # 括号里的对应发件人邮箱昵称、发件人邮箱账号
            msg['Subject'] = "供应商平台通知："  # 邮件的主题，也可以说是标题

            server = smtplib.SMTP_SSL("smtp.qq.com", 465)  # 发件人邮箱中的SMTP服务器，端口是465
            server.login(my_sender, my_pass)  # 括号中对应的是发件人邮箱账号、邮箱密码
            if len(ttx) > 5:
                my_user = my_user + ';' + ttx
            else:
                msg = MIMEText('供应商【' + r4 + '】邮箱不存在，请T100修改邮箱后，手动发送邮件给供应商登录平台</p>', 'html', 'utf-8')
            msg['To'] = formataddr(["收件人昵称", my_user])  # 括号里的对应收件人邮箱昵称、收件人邮箱账号
            server.sendmail(my_sender, [my_user, ], msg.as_string())  # 括号中对应的是发件人邮箱账号、收件人邮箱账号、发送邮件
            server.quit()  # 关闭连接
            print('email ok')

            r10 = tao[10]
            r11 = tao[11]
            r12 = tao[12]
            r19 = tao[19].strftime("%Y%m%d")
            r20 = tao[20]
            r21 = tao[21]
            r22 = tao[22]
            r23 = tao[23]
            if r23 is None:
                r23=0
            r24 = tao[24]
            r25 = tao[28]#tao[25]
            r26 = tao[26]
            r27 = tao[27]
            r28 = tao[28]
            if r28 is None:
                r28 = ''
            r29 = tao[29]
            if r29 is None:
                r29 = ''
            r30 = tao[30]
            if r30 is None:
                r30 = ''
            r23=r30.strip() + ',' + r29.strip() + ','+ r28.strip()
            print(r19,r20,r21,r22,r23,r3, r4, r5, r6, r7, r8, r9, r10,r11,r12, r0, r1, r2)
            sql = "update basedata_pmdg set pmdg008=('%s'), pmdg009=('%s'), pmdg001=('%s'),  pmaal004=('%s'), pmdg003=('%s'), imaal003=('%s'), imaal004=('%s'), pmdgua006=('%d'), pmdg007=('%f'), pmdg017=('%s'),  pmdg030=('%s'), pmdf002=('%s'), ooag011=('%s'),pmdgud017=('%d'),pmdgud005=('%s'),pmdgud018=('%f'),pmdgud019=('%f'),pmdgud020=('%f')  WHERE pmdgsite=('%s') and pmdgdocno=('%s') and  pmdgseq=('%s')  and  pmdg002=('%s')" % ( r11,r12,r3, r5, r6, r7, r8, r9,r10,r19,r20,r21,r22,r23, r24, r25, r26, r27, r0, r1, r2, r4)
            openpg.execute(sql)
            conn1.commit()  # 执行插入

            r13 = tao[13]
            r14 = tao[14]
            r15 = tao[15]
            r16 = tao[16]
            r17 = tao[17]
            r18 = tao[18]
            if r13 is None:
                r13 = 0
            if r14 is None:
                r14 = 0
            if r15 is None:
                r15 = 0
            if r16 is None:
                r16 = 0
            if r17 is None:
                r17 = 0
            if r18 is None:
                r18 = 0
            print(r13, r14, r15,r16, r17, r18)
            sql = "update basedata_pmdg set creator=pmdf002,pmdg011=('%f'), pmdgud014=('%d'), pmdg013=('%d'), pmdgua007=('%f'), pmdgud013=('%f'), pmdgud012=('%f') WHERE pmdgsite=('%s') and pmdgdocno=('%s') and  pmdgseq=('%s')  and  pmdg002=('%s')" % (r13, r14, r15,r16, r17, r18,r0, r1, r2,r4)
            openpg.execute(sql)
            conn1.commit()  # 执行插入
            print(r4)

            # sql = "select id from  auth_user  where  username=('%s')" % ( r21)
            # openpg.execute(sql)
            # rest = openpg.fetchone()
            # k2 = rest[0]
            #
            # sql = "select id from  auth_user  where  username=('%s')" % ( r4)
            # openpg.execute(sql)
            # rest = openpg.fetchone()
            # k3 = rest[0]



            # sql = "insert into workflow_instance (code, modal_id,status, object_id, starter_id,start_time) values ('N01',1,1,'%d','%d',now())" % ( k0,k2)
            # openpg.execute(sql)
            # conn1.commit()  # 执行插入
            #
            # sql = "select id from  workflow_instance  where  modal_id=1 and status=1 and object_id=('%d')" % ( k0)
            # openpg.execute(sql)
            # rest = openpg.fetchone()
            # i2 = rest[0]
            # print(i2,k2)
            # sql = "insert into workflow_history (inst_id, user_id,pro_time,pro_type) values ('%d','%d',now(),0)" % (i2,k2)
            # openpg.execute(sql)
            # conn1.commit()  # 执行插入
            # sql = "insert into workflow_todolist (inst_id,user_id,app_name,model_name,is_read,read_time,status,arrived_time) values ('%d','%d','basedata','pmdg',FALSE,now(),True,now())" % (i2,k2)
            # openpg.execute(sql)
            # conn1.commit()  # 执行插入
            #
            # sql = "insert into workflow_todolist (inst_id,user_id,app_name,model_name,node_id,is_read,arrived_time,status) values ('%d','%d','basedata','pmdg',1,FALSE ,now(),FALSE)" % (i2,k3)
            # openpg.execute(sql)
            # conn1.commit()  # 执行插入
            # sql = "insert into workflow_todolist (inst_id,user_id,app_name,model_name,node_id,is_read,arrived_time,status) values ('%d','%d','basedata','pmdg',2,FALSE ,now(),FALSE)" % (i2,k2)
            # openpg.execute(sql)
            # conn1.commit()  # 执行插入


        # except Exception:  # 如果 try 中的语句没有执行，则会执行下面的 ret=False
        #     print('email error')
        #     ret = False

    ##########从T100供应商信息表中获取代码设置为POSTGRESQL供应商平台账户,密码yxh123,完成授权并发送邮件
    xxx = "select pmaa001,pmaal004,replace(listagg(oofc012,';') within group (order by oofc012 desc),' ','') fs from  dsdata.pmaa_t left join pmaal_t on pmaa001=pmaal001 left join  dsdata.pmaj_t on pmaj001=pmaa001 and pmaalent='60'   and pmaal002='zh_CN' left join  dsdata.oofc_t on pmaj002=oofc002 and oofcent='60' and oofcstus not in ('D','R','X','E') and (oofc009='10' or oofc009 is null) and oofc012 is not null where pmaaent='60' and pmaa002 in('1','3') and oofc009='10' and oofc008='4' group by pmaa001,pmaal004"
    cur.execute(xxx)
    res = cur.fetchall()
    for tao in res:
        r0 = tao[0]
        r1 = tao[1]
        r2 = tao[2]

        sql = "select username from  auth_user  where  username=('%s')" % (r0)
        openpg.execute(sql)
        conn1.commit()  # 执行插入
        res = openpg.fetchone()
        if not res:
            print(r0, r1, r2)
            sql = "insert into auth_user (password,  is_superuser, username, first_name,  email, is_staff, is_active,last_name,date_joined) values ('pbkdf2_sha256$120000$CqYFUczpD3Pq$PUiF/M9rJ9TzEW7OIXCsXPBYS2Xs0V1Dr8Bs+thKgf8=',false ,'%s','%s','%s',true ,true,'',LOCALTIMESTAMP )" % (
            r0, r1, r2)
            openpg.execute(sql)
            conn1.commit()  # 执行插入
        sql = "select id from  auth_user  where  username=('%s')" % (r0)
        openpg.execute(sql)
        conn1.commit()  # 执行插
        res = openpg.fetchall()
        for tao in res:
            r3 = tao[0]
            try:
                sql = "INSERT INTO auth_user_groups(user_id, group_id)  VALUES ('%d',1)" % (r3)
                openpg.execute(sql)
                conn1.commit()  # 执行插入
                try:
                    my_user = 'finance@lutec.net'  # 收件人邮箱账号，我这边发送给自己
                    msg = MIMEText(
                        '亲爱的供应商【' + r1 + '】你好：<br><p>我们邀请您使用LUTEC供应商管理平台：</p><p><a href="http://www.umenb.com:8000">宁波耀泰(集团)管理平台</a><br>用户名:' + r0 + '<br>密码:yxh12345<br>警告：请登录平台及时修改默认密码，避免信息泄露<br><hr>技术支持：宁波耀泰(集团)，<a href=mailto:finance@lutec.net>客服邮箱</a></p>',
                        'html', 'utf-8')
                    msg['From'] = formataddr(["发件人昵称", my_sender])  # 括号里的对应发件人邮箱昵称、发件人邮箱账号
                    msg['Subject'] = "LUTEC供应商平台通知："  # 邮件的主题，也可以说是标题

                    server = smtplib.SMTP_SSL("smtp.qq.com", 465)  # 发件人邮箱中的SMTP服务器，端口是465
                    server.login(my_sender, my_pass)  # 括号中对应的是发件人邮箱账号、邮箱密码
                    if len(r2) > 5:
                        my_user = my_user + ';' + r2
                    else:
                        msg = MIMEText('供应商【' + r1 + '】邮箱不存在，请T100修改邮箱后，手动发送邮件给供应商登录平台</p>', 'html', 'utf-8')
                    msg['To'] = formataddr(["收件人昵称", my_user])  # 括号里的对应收件人邮箱昵称、收件人邮箱账号
                    server.sendmail(my_sender, [my_user, ], msg.as_string())  # 括号中对应的是发件人邮箱账号、收件人邮箱账号、发送邮件
                    server.quit()  # 关闭连接
                    print(r0, r1, r3)
                except Exception:  # 如果 try 中的语句没有执行，则会执行下面的 ret=False
                    ret = False
            except Exception:
                conn1.rollback()
                #print(r0, r1,r2,r3)

    ##########apca_t付款计划

    sql = "UPDATE apca_t Set ok=true "
    openpg.execute(sql)
    conn1.commit()  # 执行插入
    xxx="select distinct  case when to_char(b.c+ooib014,'dd')<='05' then to_char(b.c+ooib014,'yyyymm')||'05'  when to_char(b.c+ooib014,'dd')<='15' then to_char(b.c+ooib014,'yyyymm')||'15'  when to_char(b.c+ooib014,'dd')<='25' then to_char(b.c+ooib014,'yyyymm')||'25' else to_char(add_months(b.c+ooib014,1),'yyyymm')||'05' end 计划付款日,  apca004 供方编号,  decode(apca004,'98000',apbb049,pmaal004) 供应商简称,  apcadocno 应付单号,  ooibl004 付款条件,  nvl(b.b,apca066) 发票号码,  b.c 发票日期,  (case when substr(apcadocno,4,2) in('75') then -1 else 1 end)*(apca108-nvl(a.b,0)) 应付预付金额,  apca053 备注,  apca040 冻结状态,  ooib014 账期天数,  apcasite 公司别,  ooag001 采购员编码,ooag011 采购员  from  DSdata.apca_t   left join (select apccdocno a,sum(apcc109) b from  DSdata.apcc_t where apccent='60'  group by apccdocno) a on a.a=apcadocno  left join (select isamdocno a,listagg(isam010,'|') within group (order by isam010) b,max(isam011) c from  DSdata.isam_t where isament='60' group by isamdocno) b on b.a=nvl(apca018,apcadocno)  left join dsdata.apbb_t on apbbent='60' and apbbcomp=apcasite and apca018=apbbdocno  left join DSdata.pmaal_t on apca004=pmaal001 and pmaalent='60'  left join (select distinct apce003 a from DSdata.apce_t left join DSdata.apda_t  on apdadocno=apcedocno and apdaent='60'  where  apdastus in('N','W') and apceent='60' and substr(apcedocno,4,2)='76') d on d.a=apcadocno  left join DSdata.pmab_t on pmabent='60' and pmabsite=apcasite and pmab001=apca004  left join DSdata.ooag_t on ooagent='60' and ooag001=apca014 and ooag004=apcasite  left join DSdata.ooib_t on pmab037=ooib002 and ooibent='60'   left join DSdata.ooibl_t on pmab037=ooibl002 and ooiblent='60'   where apcaent='60'  and apcastus in('Y','A') and (apca108-nvl(a.b,0))>0 and substr(apcadocno,4,3) in('710','711','715','720') and substr(apca008,1,2) not in('11') and apca004<>'98000'"
    cur.execute(xxx)
    res = cur.fetchall()
    for tao in res:
        r0 = tao[0]#.strftime("%Y%m%d")
        r1 = tao[1]
        r2 = tao[2]
        r3 = tao[3]
        try:
            sql = "select paybill from apca_t where paybill=('%s')" % (r3)
            openpg.execute(sql)
            conn1.commit()  # 执行插入
            res = openpg.fetchone()
            if not res:
                sql = "insert into apca_t (paybill) values ('%s')" % (r3)
                openpg.execute(sql)
                conn1.commit()  # 执行插入
                r4 = tao[4]
                r5 = tao[5]
                r6 = tao[6]
                r7 = tao[7]
                r8 = tao[8]
                r9 = tao[9]
                r10 = tao[10]
                r11 = tao[11]
                r12 = tao[12]
                r13 = tao[13]
                try:
                    r14 = r0[0:4] + '.' + datetime.datetime.strptime(r0,'%Y%m%d').strftime("%W")
                except Exception:
                    r14 = ''
                if r5 is None:
                    r5=''
                if r6 is None:
                    r6=''
                if r7 is None:
                    r7=0
                if r8 is None:
                    r8=''
                if r9 is None:
                    r9=''
                if r10 is None:
                    r10=0
                if r11=='Y1':
                    r11='Y1-耀泰'
                if r11=='Y3':
                    r11='Y3-颐道'
                sql = "UPDATE apca_t Set paydate=('%s'),sno=('%s'),  sname=('%s'), paytj=('%s'),fphm=('%s'),fprq=('%s'),cash=('%d'),note=('%s'),  status=('%s'), zlts=('%d'),gsb=('%s'),cgybm=('%s'),  cgy=('%s'),ok=FALSE,zc=('%s')  WHERE paybill=('%s') " % (
                    r0, r1, r2, r4, r5, r6, r7, r8, r9, r10, r11, r12, r13,r14, r3)
                try:
                    openpg.execute(sql)
                    conn1.commit()  # 执行插入
                except Exception:
                    print(r0, r1, r2, r4, r5, r6, r7, r8, r9, r10, r11, r12, r13, r14, r3)
            sql = "UPDATE apca_t Set ok=FALSE  WHERE paybill=('%s') " % ( r3)
            try:
                openpg.execute(sql)
                conn1.commit()  # 执行插入
            except Exception:
                print(r0, r1, r2, r4, r5, r6, r7, r8, r9, r10, r11, r12, r13, r14, r3)
        except Exception:
            print('error', r3)
    ##########apba_t财务对账单
    try:
        sql = "UPDATE apba_t Set ok=true "
        openpg.execute(sql)
        conn1.commit()  # 执行插入
    except Exception:
        conn1.rollback()
        print('update error')
    xxx="select apbb002 供应商编号,decode(apbb002 ,'98000',apbb049,pmaal004) 供应商简称,apbadocno 对账单号,apbaseq 项次,pmds001 进货日期,pmdt001 采购单号,apba007||':'||imaal003 货物劳务名称, imaal004 规格型号, apba012*apba010 数量, apba014 单价, apba009 单位, apba012*apba103 金额, apba012*apba104 税额, apba012*apba105 合计, pmdtdocno||' '||pmdtseq 进货单号, pmdtud003 订单号, pmdt059 备注, apbbcomp 公司别,apbbstus from dsdata.apba_t left join dsdata.apbb_t  on apbbdocno=apbadocno  and apbbent=apbaent left join dsdata.imaa_t  on apba007=imaa001 and imaaent=apbaent left join dsdata.imaal_t  on imaa001=imaal001 and imaalent='60' and imaal002='zh_CN' left join dsdata.pmdt_t  on pmdtdocno=apba005 and pmdtseq=apba006 and pmdtent=apbaent left join dsdata.pmds_t  on pmdsdocno=pmdtdocno  and pmdsent=apbaent left join DSdata.pmaal_t on apbb002=pmaal001 and pmaalent=apbaent where apbaent='60' and apbbstus='N' and apbb002<>'98000'"
    cur.execute(xxx)
    res = cur.fetchall()
    for tao in res:
        r0 = tao[0]
        r1 = tao[1]
        r2 = tao[2]
        r3 = tao[3]
        sql = "select paybill from  apba_t  where  paybill=('%s') and xc=('%s')" % (r2,r3)
        openpg.execute(sql)
        res = openpg.fetchone()
        if not res:
            sql = "insert into apba_t (paybill,xc) values ('%s','%s')" % (r2,r3)
            openpg.execute(sql)
            conn1.commit()  # 执行插入
        r4 = tao[4].strftime("%Y%m%d")
        r19 = tao[4].strftime('%Y%m%d')[0:4] + '.' + tao[4].strftime("%W")
        r5 = tao[5]
        r6 = tao[6]
        r7 = tao[7]
        r8 = tao[8]
        r9 = tao[9]
        r10 = tao[10]
        r11 = tao[11]
        r12 = tao[12]
        r13 = tao[13]
        r14 = tao[14]
        r15 = tao[15]
        r16 = tao[16]
        r17 = tao[17]
        r18 = tao[18]

        if r5 is None:
            r5=''
        if r6 is None:
            r6=''
        if r7 is None:
            r7=''
        if r9 is None:
            r9=0
        if r11 is None:
            r11=0
        if r12 is None:
            r12=0
        if r14 is None:
            r14=''
        if r15 is None:
            r15=''
        if r16 is None:
            r16=''
        if r6 is not None:
            r6 = r6.replace("'", "")
        if r7 is not None:
            r7 = r7.replace("'", "")
        if r17=='Y1':
            r17='Y1-耀泰'
        if r17=='Y3':
            r17='Y3-颐道'
        r14 = r14.replace("'", "")
        r16 = r16.replace("'", "")

        sql = "UPDATE apba_t Set sno=('%s'),sname=('%s'),  jhrq=('%s'), cgdh=('%s'),lwmc=('%s'),ggxh=('%s'),sl=('%d'),price=('%d'),  dw=('%s'), cash=('%d'),tax=('%d'),slv=0.13,totalcash=('%d'),jhdh=('%s'),  ddhm=('%s'), note=('%s'),gsb=('%s'),apbbstus=('%s'),ok=false,zc=('%s') WHERE paybill=('%s') and xc=('%s')"  % (
         r0, r1, r4, r5, r6,r7,r8, r9, r10, r11,r12, r13, r14, r15, r16,r17,r18,r19,r2, r3)
        try :
            openpg.execute(sql)
            conn1.commit()  # 执行插入
        except Exception:
            print( r14, r15, r16,r17,r18,r19,r2, r3)
    ##########Pmdnt200
    sql = "UPDATE \"Pmdnt200\" Set ok=true "
    openpg.execute(sql)
    conn1.commit()  # 执行插入

    xxx="SELECT  供方编号, 简称, 全称, 采购单号序号,采购单别,采购单号, 序号, 采购日期,  品号, 品名, 规格, 计量单位, 采购量, 已交数量, 未交量,  币别, 采购单价, 税率, 未交未税金额, 税额, 未交含税金额, 分类, 预交日, 来源单号,   PO, 采购人员, 审核者, tao料号, 公司 FROM yxhst20190325 "

    xxx="SELECT pmaal001 供方编号,pmaal004 简称,pmaal004 全称,pmdndocno||'-'||pmdnseq 采购单号序号,substr(pmdndocno,4,4) 采购单别,substr(pmdndocno,9,10) 采购单号,pmdnseq 序号,pmdldocdt 采购日期,pmdn001 品号,imaal003 品名,imaal004 规格,pmdn010 计量单位,pmdn007 采购量,a.qty 已交数量,pmdn007-nvl(a.qty,0) 未交量,pmdl015 币别,pmdn015 采购单价,pmdn017 税率,round(pmdn015*(pmdn007-nvl(a.qty,0)),2) 未交未税金额,round(pmdn015*(pmdn007-nvl(a.qty,0))*pmdn017/100,2) 税额,round(pmdn015*(pmdn007-nvl(a.qty,0)),2)+round(pmdn015*(pmdn007-nvl(a.qty,0))*pmdn017/100,2) 未交含税金额,decode(pmdl005,1,'采购','加工') 分类,pmdn012 预交日,pmdnua002||'-'||pmdnua003 来源单号,xmda033  PO,a1.ooag011 采购人员,a1.ooag011 审核者,regexp_substr(imaal003,'[A-Z]{2,4}\d{4,5}[A-Z]*|\d{8,10}')  tao料号,pmdlsite 公司,Nvl(pmdl044,'')||Nvl(pmdn050,'') 备注,pmdl002 采购员编码 FROM dsdata.pmdl_t left join dsdata.pmdn_t on pmdnent=pmdlent and pmdndocno=pmdldocno left join dsdata.imaa_t  on imaaent=pmdlent  and imaa010<>'SX' and  pmdn001=imaa001 left join dsdata.imaal_t  on imaa001=imaal001  and imaalent='60' left join dsdata.pmaa_t on pmaaent=pmdlent and pmaa002 in('1','3') and pmaa001=pmdl004 left join dsdata.pmaal_t on pmaalent=pmdlent  and pmaal001=pmaa001 left join (select pmdodocno||pmdoseq no,sum(pmdo019) qty    from dsdata.pmdl_t left join dsdata.pmdo_t  on pmdodocno=pmdldocno and  pmdoent=pmdlent   where pmdlent='60' and pmdlstus in('A','F','Y')     group by pmdodocno||pmdoseq)  a on a.no=pmdndocno||pmdnseq left join dsdata.ooag_t a1 on a1.ooagent=pmdlent and a1.ooag001=pmdl002  left join dsdata.ooag_t a2 on a2.ooagent=pmdlent and a2.ooag001=pmdlcnfid left join dsdata.xmda_t on pmdnua002=xmdadocno and xmdaent=pmdlent where pmdlent='60' and pmdlstus in('A','F','Y','N') and pmdn045='1' and pmdn007-nvl(a.qty,0)>0 and substr(pmdndocno,4,4)<>'3389'  and pmaal001<>'98000' and substr(pmdndocno,4,4)<>'3399' order by pmaal001||'|'||pmaal004,pmdn012,pmdndocno"
    cur.execute(xxx)
    res = cur.fetchall()

    for tao in res:
        r0 = tao[0]
        r1 = tao[1]
        r2 = tao[2]
        r3 = tao[3]
        sql = "select billno from  \"Pmdnt200\"  where  billno=('%s')" % (r3)
        openpg.execute(sql)
        res = openpg.fetchone()
        if not res:
            sql = "insert into \"Pmdnt200\" (billno) values ('%s')" % (r3)
            openpg.execute(sql)
            conn1.commit()  # 执行插入
        r4 = tao[4]
        r5 = tao[5]
        r6 = tao[6]
        r7 = tao[7]
        r8 = tao[8]
        r9 = tao[9]
        r10 = tao[10]
        if r9 is not None:
            r9 = r9.replace("'", "")
        if r10 is not None:
            r10 = tao[10].replace("'", "")
        r11 = tao[11]
        r12 = tao[12]
        sql = "UPDATE \"Pmdnt200\" Set sno=('%s'), jc=('%s'), allname=('%s'), bclass=('%s'),bbill=('%s'),ok=false  WHERE billno=('%s')" % (r0,r1,r2,r4,r5,r3)
        openpg.execute(sql)
        conn1.commit()  # 执行插入
        sql = "UPDATE \"Pmdnt200\" Set bnum=('%d'),buydate=('%s'),  code=('%s'), name=('%s'),spec=('%s'),unit=('%s')  WHERE billno=('%s')" % (
         r6, r7, r8, r9, r10, r11, r3)
        openpg.execute(sql)
        conn1.commit()  # 执行插入
        r13 = tao[13]
        r14 = tao[14]
        r15 = tao[15]
        r16 = tao[16]
        r17 = tao[17]
        r18 = tao[18]
        r19 = tao[19]
        r20 = tao[20]
        r21 = tao[21]
        r22 = tao[22]
        r23 = tao[23]
        r24 = tao[24]
        r25 = tao[25]
        r26 = tao[26]
        r27 = tao[27]
        r28 = tao[28]
        if r28=='Y1':
            r28='Y1-耀泰'
        if r28=='Y3':
            r28='Y3-颐道'
        r29 = tao[29]
        r30 = tao[30]
        if r18 is None:
            r18=0
        if r19 is None:
            r19=0
        if r20 is None:
            r20=0
        if r13 is None:
            r13=0
        if r22 is None:
            sql = "UPDATE \"Pmdnt200\" Set notex=('%d'),tex=('%d'),  allchash=('%d'), sclass=('%s'),sbill=('%s'),note=('%s'),ok=false  WHERE billno=('%s')" % (
                r18, r19, r20, r21,  r23, r29,r3)
        else:
            sql = "UPDATE \"Pmdnt200\" Set notex=('%d'),tex=('%d'),  allchash=('%d'), sclass=('%s'),predate=('%s'),sbill=('%s')  WHERE billno=('%s')" % (
                r18, r19, r20, r21, r22, r23, r3)

        openpg.execute(sql)
        conn1.commit()  # 执行插入
        if r12 is None:
            r12=0
        if r14 is None:
            r14=0
        if r15 is None:
            r15=0
        if r16 is None:
            r16=0
        if r17 is None:
            r17=0
        if r22 is not None:
            r29 = r22.strftime('%Y-%m-%d')[0:4] + '.' + r22.strftime("%W")
        else:
            r29 = ''


        sql = "UPDATE \"Pmdnt200\" Set quan=('%d'),okquan=('%d'),  noquan=('%d'), currenct=('%s'),price=('%d'),rate=('%d'),zc=('%s')  WHERE billno=('%s')" % (
         r12, r13, r14, r15, r16, r17,r29, r3)
        openpg.execute(sql)
        sql = "UPDATE \"Pmdnt200\" Set \"PO\"=('%s'),buyer=('%s'),  checker=('%s'), tao=('%s'),company=('%s'),buyercode=('%s') WHERE billno=('%s')" % (
         r24, r25, r26, r27, r28,r30,r3)
        openpg.execute(sql)
        conn1.commit()  # 执行插入
    conn1.close
    conn2.close
    print('OK')
        # sql = "UPDATE \"Pmdnt200\" Set sno=('%s'), jc=('%s'), allname=('%s'), bclass=('%s'),bbill=('%s'),bnum=('%s'),buydate=('%s'),  code=('%s'), name=('%s'),spec=('%s'),unit=('%s'), '采购量'=('%d'), '已交数量'=('%d'), '未交量'=('%d'), '供方回复交期'=('%s'), '供方说明'=('%s'), '币别'=('%s'), '采购单价'=('%d'), '税率'=('%d'), '未交未税金额'=('%d'), '税额'=('%d'), '未交含税金额'=('%d'), '分类'=('%s'), '预交日'=('%s'), '来源单号'=('%s'), '备注'=('%s'),  'PO'=('%s'), '采购人员'=('%s'), '审核者'=('%s'), 'tao料号'=('%s'), '公司'=('%s')  WHERE '采购单号序号'=('%s')" % (r0,r1,r2,r4,r5,r6,r7,r8,r9,r10,r11,r12,r13,r14,r15,r16,r17,r18,r19,r20,r21,r22,r23,r24,r25,r26,r27,r28,r29,r30,r31,r3)
        # sql = "UPDATE Pmdnt100 Set '供方编号'=('%s'), '简称'=('%s'), '全称'=('%s'), '采购单别'=('%s'),'采购单号'=('%s'), '序号'=('%s'), '采购日期'=('%s'), '供方简称'=('%s'), '品号'=('%s'), '品名'=('%s'), '规格'=('%s'), '计量单位'=('%s'), '采购量'=('%d'), '已交数量'=('%d'), '未交量'=('%d'), '供方回复交期'=('%s'), '供方说明'=('%s'), '币别'=('%s'), '采购单价'=('%d'), '税率'=('%d'), '未交未税金额'=('%d'), '税额'=('%d'), '未交含税金额'=('%d'), '分类'=('%s'), '预交日'=('%s'), '来源单号'=('%s'), '备注'=('%s'),  'PO'=('%s'), '采购人员'=('%s'), '审核者'=('%s'), 'tao料号'=('%s'), '公司'=('%s')  WHERE '采购单号序号'=('%s')" % (r0,r1,r2,r4,r5,r6,r7,r8,r9,r10,r11,r12,r13,r14,r15,r16,r17,r18,r19,r20,r21,r22,r23,r24,r25,r26,r27,r28,r29,r30,r31,r3)
        # cr.execute(sql)
        # conn.commit()  # 执行插入

def get_sale_detail():
    conn = psycopg2.connect("host=%s port=%s dbname=%s user=%s password=%s" % (DBHOST, DBPORT, DBNAME, DBUSER, DBPASS))
    cr = conn.cursor()
    conn1 = pymssql.connect(host='192.168.0.2', user='sa', password='yh***microsoft***', database='trade',
                            charset='utf8')
    cur1 = conn1.cursor()

    conn2 = cx_Oracle.connect('dsdata/dsdata@192.168.0.5:1521/TOPPRD')
    cur = conn2.cursor()

    #sheet_1 = json.dumps(data)
    cr.execute("SELECT distinct sale_order.id,res_partner.ref FROM sale_order left join res_partner on res_partner.id=sale_order.partner_id  order by sale_order.id desc limit 10")
    res = cr.fetchall()
    for tao in res:
        row = tao[0]
        customer = tao[1]
        newFile = open("d:/tao_order/tmp.xlsx", "wb")
        oldFile = open("d:/tao_order/T100.xlsx", "rb")
        contents = oldFile.readlines()
        newFile.writelines(contents)
        newFile.close()
        oldFile.close()

        workbook = openpyxl.load_workbook(r"d:/tao_order/tmp.xlsx")
        worksheet = workbook.worksheets[0]

        row0 = 'd:/tao_order/'+str(row)+'.xlsx'
        f = xlwt.Workbook()  # 创建工作簿
        print(">>> CONNECTION OK", customer)
        if os.path.isfile(row0) == False:
            cr.execute("SELECT partner_id,date_order,item_number,sale_order_line.name,product_uom_qty,price_reduce,sale_order.name,client_order_ref ,note,crd_date,psi_date,requested_date,commitment_date,res_currency.name,sale_order_line.product_id FROM sale_order left join sale_order_line on sale_order.id=sale_order_line.order_id  left join res_currency on res_currency.id=sale_order_line.currency_id where sale_order.id=%d order by item_number" % (row))
            res = cr.fetchall()
            rownum=2
            for row in res:
                partner_id = row[0]
                date_order = row[1]
                item_number = row[2]
                product_name = str(row[3])
                product_uom_qty = row[4]
                price_reduce = row[5]
                name = row[6]
                client_order_ref = row[7]
                note = row[8]
                crd_date = row[9]
                psi_date = row[10]
                requested_date = row[11]
                commitment_date = row[12]
                currency_id = row[13]
                product_id = str(row[14])
                if date_order == None:
                    date_order = ''
                else:
                    date_order = date_order + datetime.timedelta(hours=8)
                    date_order = date_order.strftime("%Y%m%d")

                if crd_date == None:
                    crd_date = ''
                else:
                    crd_date = crd_date + datetime.timedelta(hours=8)
                    crd_date = crd_date.strftime("%Y%m%d")
                if psi_date == None:
                    psi_date=''
                else:
                    psi_date = psi_date + datetime.timedelta(hours=8)
                    psi_date=psi_date.strftime("%Y%m%d")
                if requested_date == None:
                    requested_date=''
                else:
                    requested_date = requested_date + datetime.timedelta(hours=8)
                    requested_date = requested_date.strftime("%Y%m%d")
                if commitment_date == None:
                    commitment_date = ''
                else:
                    commitment_date = commitment_date + datetime.timedelta(hours=8)
                    commitment_date = commitment_date.strftime("%Y%m%d")
                if client_order_ref == None:
                    client_order_ref=''
                namet100 = name + client_order_ref
                cur.execute("SELECT pmaal001 FROM pmaal_t where pmaal005='%s'" % (customer))
                row = cur.fetchone()
                if row == None:
                    partner_id = product_name
                    print('lhb1:', customer, partner_id)
                else:
                    partner_id=str(row[0])
                    print('lhb2:',customer,partner_id)
                product_name = product_name[1:9]

                cur.execute("SELECT imaa001 FROM imaa_t where imaaua613='%s'" % (product_name))
                row = cur.fetchone()
                if row == None:
                    #product_id = partner_id
                    print('lhb1:', customer, product_id)
                else:
                    product_id=row[0]
                    print('lhb2:',customer,product_id)

                # 通过get_sheet()获取的sheet有write()方法
                worksheet.cell(rownum, 3, partner_id)
                worksheet.cell(rownum, 4, date_order)
                worksheet.cell(rownum, 5, date_order)
                worksheet.cell(rownum, 6, item_number)
                worksheet.cell(rownum, 7, product_id)
                worksheet.cell(rownum, 10, product_uom_qty)
                worksheet.cell(rownum, 11, 100)
                worksheet.cell(rownum, 12, price_reduce)
                worksheet.cell(rownum, 13, namet100)
                worksheet.cell(rownum, 14,  'N')
                worksheet.cell(rownum, 15, note)
                worksheet.cell(rownum, 16, crd_date)
                worksheet.cell(rownum, 17, psi_date)
                worksheet.cell(rownum, 18,requested_date)
                worksheet.cell(rownum, 19, commitment_date)
                worksheet.cell(rownum, 20, 'Y00028')
                worksheet.cell(rownum, 21,currency_id)
                rownum = rownum+1

            workbook.save(filename=row0)
                #if sendname=='陆佳仪' or sendname=='姚旭辉' or sendname=='Coco女儿': # 你朋友的微信名称，不是备注，也不是微信帐号。
            time.sleep(2)
            #my_friend = bot.groups().search('Tao订单释放信息')[0]
            time.sleep(2)
                #else:
                #    my_friend = bot.groups().search(sendname)[0]
            #my_friend.send_file(row0)
            # time.sleep(2)
            # my_friend = bot.friends().search('丹丹')[0]
            # my_friend.send_file(row0)
            # # 发送文本
            # my_friend.send('Hello, WeChat!')
            # # 发送图片
            # my_friend.send_image('my_picture.png')
            # # 发送视频
            # my_friend.send_video('my_video.mov')
            # # 发送文件
            # my_friend.send_file('my_file.zip')
            # # 以动态的方式发送图片
            # my_friend.send('@img@my_picture.png')
    if  datetime.datetime.now().hour == 16:
        nowdate = datetime.datetime.now().strftime('%Y%m%d')
        sql = "select name FROM [getsmm] where CONVERT(varchar(100),[creatdate], 112) = ('%s')"% (nowdate)
        cur1.execute(sql)
        resone = cur1.fetchone()
        if resone == None:
            content = '爬取SMM，没有记录，请手动保存原网页'
            #my_friend = bot.friends().search('姚旭辉')[0]
            #my_friend.send(content)

    cur1.close
    conn1.close
    conn2.close
def tao_sale_order_bop():
    conn = psycopg2.connect("host=%s port=%s dbname=%s user=%s password=%s" % (DBHOST, DBPORT, DBNAME, DBUSER, DBPASS))
    cr = conn.cursor()
    conn1 = pymssql.connect(host='192.168.0.2', user='sa', password='yh***microsoft***', database='trade',
                            charset='utf8')
    cur1 = conn1.cursor()

    conn2 = cx_Oracle.connect('dsdata/dsdata@192.168.0.5:1521/TOPPRD')
    cur = conn2.cursor()

    print(">>> START", datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
    print(">>> CONNECTION OK")
    get_sale_detail()
    sql = "update [tao_sale_order_bop] set id=2 where id < 2"
    cur1.execute(sql)
    conn1.commit()  # 执行插入
    #cr.execute("SELECT item, name, description, status, type, mfg, scrap_rate, bom_unit, purchase_unit, write_date FROM lutec_p#roduct")
    cr.execute("select name FROM sale_order where pack_approval_date is null order by pack_approval_date")
    res = cr.fetchall()
    i = 1
    for name in res:
        try:
            i = i + 1
            name = name[0]
            #print("> READ Product: %d, %s,%s" % (i,name,bop_approval_date))
            sql = "select name from [tao_sale_order_bop] where [name] ='%s'" % (name)
            cur1.execute(sql)
            resone = cur1.fetchone()
            #     #conn.commit()  # 执行插入
            # except Exception:
            #     print("> READ tao_sale_order_bop: %d, %s" % (i, name))
            if resone==None:
                #resone = tsn.fetchone()
                sql = "insert into [tao_sale_order_bop] ([name])  values ('%s')" % (name)

            else:
                sql = "update [tao_sale_order_bop] set creatdate=getdate(),id=1 where [name] = ('%s')" % (name)
            try:
                cur1.execute(sql)
                conn1.commit()  # 执行插入
            except Exception:
                print("%s" % (traceback.format_exc()))
        except Exception as err:
            print("%s" % (traceback.format_exc()))
    sql = "select name from [tao_sale_order_bop] where id = 2"
    yy = cur1.execute(sql)
    res=cur1.fetchall()
    if res==None:
        print("> READ yy: None" )
    else:
            #res = yy.fetchall()
            for name in res:
                namec = name[0]
                xx = cr.execute("select pack_approval_date,state FROM sale_order where name='%s'" % (namec))
                rowstate = cr.fetchone()
                try:
                    if rowstate == None:
                        row0='空值'
                        row1 = '空值'
                    else:
                        row0 = rowstate[0].strftime("%Y-%m-%d %H:%M:%S")

                        row1 = rowstate[0]+datetime.timedelta(hours=8)
                        row1 = row1.strftime("%Y-%m-%d %H:%M:%S")
                    try:
                        rows = rowstate[1]
                    except Exception:
                        rows = '空值'
                    print("> READ tao_sale_order_bop: %s, %s" % (row0, row1))
                    try:
                        content = 'TAOUP订单BOP:'+namec+',bop_approval_date：'
                        content = content+row1+',state:'
                        content = content+rows
                        maxid=maxinterid('rtxmessage')
                        sql = "UPDATE tablemaxid Set id=%d  WHERE tablename='rtxmessage'" % (maxid)
                        cur1.execute(sql)
                        conn1.commit()  # 执行插入
                        sql = "insert into [rtxmessage] (interid) values ('%d')" % (maxid)
                        try:
                            cur1.execute(sql)
                            sql = "UPDATE [rtxmessage]  SET note=%s, toman='????;????;',billname='????',title='TAOUP订单BOP',sysid= 0 WHERE interid=%d"
                            cur1.execute(sql,(content,maxid))
                            conn1.commit()  # 执行插入
                            sql = "UPDATE [tao_sale_order_bop] SET id=9 WHERE name='%s'" % (namec)
                            cur1.execute(sql)
                            conn1.commit()  # 执行插入
                        except Exception:
                            print("> READ rtxmessage: %s, %s" % (row0, row1))
                            print("%s" % (traceback.format_exc()))
                    except Exception:
                        print("> READ tablemaxid: %s, %s" % (row0, row1))
                        print("%s" % (traceback.format_exc()))
                except Exception:
                    print("> READ tao_sale_order_bop: %s",namec)
                    print("%s" % (traceback.format_exc()))
    cur1.close
    conn1.close
    conn2.close
    get_t100()

def copyuser():
    import cx_Oracle
    DBNAME1 = 'CIMS'
    DBHOST1 = 'localhost'
    DBUSER1 = 'openpg'
    DBPASS1 = 'xxx'
    DBPORT1 = 5432
    conn1 = psycopg2.connect("host=%s port=%s dbname=%s user=%s password=%s" % (DBHOST1, DBPORT1, DBNAME1, DBUSER1, DBPASS1))
    openpg = conn1.cursor()
    conn2 = cx_Oracle.connect('dsdata/dsdata@192.168.0.5:1521/TOPPRD')
    cur = conn2.cursor()

    workbook = xlrd.open_workbook("D:\code.xlsx")
    sheet = workbook.sheet_by_name('Sheet1')
    nrows = sheet.nrows
    for i in range(nrows):
        textList = sheet.row_values(i)
        code = textList[0]
        name = textList[1]
        passw=textList[2]+textList[1]
        passw=hashlib.sha1(passw.encode("utf8")).hexdigest()
        USERT = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        sql = "insert into public.auth_user (username, first_name,password,is_superuser, is_staff, is_active, date_joined,last_name,email) values ('%s','%s','%s',FALSE,TRUE ,TRUE,'%s','','')" % (code,name,passw,USERT)
        openpg.execute(sql)
    conn1.commit()  # 执行插入

conn2 = pymssql.connect(host='192.168.0.2', user='sa', password='yh***microsoft***', database='trade',charset='utf8')
cur2 = conn2.cursor();

if not cur2:
    raise Exception('数据库连接失败！')

def maxinterid(table1):
    sql = "select id from [tablemaxid] where tablename=%s"
    cur2.execute(sql, (table1))
    row = cur2.fetchone()
    try:
        row0 = row[0]
    except Exception:
        row0 = 1
    today=datetime.datetime.now()
    maxid=today.year*1000000+today.month*10000
    if maxid<=row0:
        row0 += 1
    else:
        row0 = maxid + 1
    # sql = "UPDATE tablemaxid Set id=%d  WHERE tablename=%s" % (maxid,table1)
    #
    # cur2.execute(sql)
    # conn2.commit()  # 执行插入
    cur2.close
    return row0
def spider_tao():
    try:
        # re2 = requests.get("https://www.qichacha.com/", headers=headersqichacha)
        # if re2.status_code == 200:
        #     f = open("D:/qcceveryday/"+  datetime.datetime.now().strftime("%Y%m%d-%H-%M"), "w", encoding='utf-8-sig')
        #     f.write(re2.text)
        # print('TAOUP工单开始', datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
        # os.system("python D:/ljyqcc/tao_mrp_production.py")
        # print('TAOUP发货开始', datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
        # os.system("python D:/ljyqcc/tao_stock_picking.py")
        print('TAOUP订单开始')
        p =os.system("python D:/ljyqcc/tao_sale_order.py")
        print('订单爬完', datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
        if os.path.isfile('D:/ljyqcc/tmpsale') == True:
            if os.path.getsize('D:/ljyqcc/tmpsale') > 10:
                with open('D:/ljyqcc/tmpsale', 'r', encoding='utf-8') as f:
                    content=f.read()
                    content='TAOUP订单：'+content
                #my_friend = bot.friends().search('姚旭辉')[0]
                #my_friend.send(content)
                f.close()
                os.remove('D:/ljyqcc/tmpsale')
        print('订单微信爬完', datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'))

    except:
        print('TAOUP失败',datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
def mini2weixin():
    # 设置配置信息
    appid = "xxx"  # 小程序appid
    secret = "xxx"  # 小程序secret
    sn = 25  # 每次推送数量
    timer = 5  # 每次睡眠时间

    # 获取小程序assess_token
    tokenUrl = "https://api.weixin.qq.com/cgi-bin/token?grant_type=client_credential&appid={}&secret={}".format(appid,secret)
    tokenReq = requests.get(tokenUrl)
    tokenResp = tokenReq.json()
    token = tokenResp['access_token']

    # 拼接推送链接
    url = "https://api.weixin.qq.com/wxa/search/wxaapi_submitpages?access_token={}".format(token)

    # 获取小程序招工推送列表
    postData = {
        "access_token": token,
        "pages": []
    }

    # 读取所有待推送信息
    with open('d:\ids.txt', 'r') as f:
        ids = f.read()
        ids = ids.split(",")  # 切割成数组保存
        idsLen = len(ids)  # 获取数组长度
        maxGroup = math.ceil(idsLen / sn)  # 最大的分组数量
        group_m = -1
        lists = []  # 使用新数组保存
        for i in range(idsLen):
            if i % sn == 0:
                group_m += 1
                lists.append([ids[i]])
            else:
                lists[group_m].append(ids[i])

        # 分组推送
        sign = 0  # 标记当前推送条数
        for item in lists:
            arrData = []  # 声明或重置待提交数组
            for i in item:
                data = {
                    "path": "pages/goods/goods",
                    "query": "id=" + i
                }
                arrData.append(data)

            postData['pages'] = arrData
            onceReq = requests.post(url, json.dumps(postData))
            onceRes = onceReq.json()
            signStart = sign * sn
            signEnd = (sign * sn) + sn
            if onceRes['errcode'] == 0:
                print(datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
                # print("当前推送第{} - {} 条：成功！最后一条数据为：{}".format(signStart, signEnd, arrData[-1]))
            elif onceRes['errcode'] == 47006:
                print(datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
                print("当前推送第{} - {} 条：失败！返回状态码：{},最后一条数据为：{}, 当日推送已达到最大上限！".format(signStart, signEnd, onceRes['errcode'],arrData[-1]))
                break
            else:
                print("当前推送第{} - {} 条：失败！返回状态码：{},最后一条数据为：{}".format(signStart, signEnd, onceRes['errcode'], arrData[-1]))

            sign += 1
            time.sleep(timer)  # 设置睡眠时间
def get_msg():
    url = 'http://open.iciba.com/dsapi/'
    response = requests.post(url).text
    data = json.loads(response)
    img = data['fenxiang_img']
    name = str(data['dateline']) + '.jpg'
    with open(name, 'wb') as f:
        f.write(requests.get(img, timeout=30).content)
    f.close()
    return name


def setImage(imgpath):
    im = Image.open(imgpath)
    im.save('1.bmp')
    aString = windll.user32.LoadImageW(0, r"1.bmp", win32con.IMAGE_BITMAP, 0, 0, win32con.LR_LOADFROMFILE)

    if aString != 0:  ## 由于图片编码问题  图片载入失败的话  aString 就等于0
        w.OpenClipboard()
        w.EmptyClipboard()
        w.SetClipboardData(win32con.CF_BITMAP, aString)
        w.CloseClipboard()
def getText():
    """获取剪贴板文本"""
    w.OpenClipboard()
    d = w.GetClipboardData(win32con.CF_UNICODETEXT)
    w.CloseClipboard()
    return d


def setText(aString):
    """设置剪贴板文本"""
    w.OpenClipboard()
    w.EmptyClipboard()
    w.SetClipboardData(win32con.CF_UNICODETEXT, aString)
    w.CloseClipboard()


def send_qq(to_who, msg):
    """发送qq消息
    to_who：qq消息接收人
    msg：需要发送的消息
    """
    # 将消息写到剪贴板
    setText(msg)
    #     sendByUser('poplar')
    # 获取qq窗口句柄
    qq = win32gui.FindWindow(None, to_who)

    # 投递剪贴板消息到QQ窗体
    win32gui.SendMessage(qq, 258, 22, 2080193)
    win32gui.SendMessage(qq, 770, 0, 0)
    # 模拟按下回车键
    win32gui.SendMessage(qq, win32con.WM_KEYDOWN, win32con.VK_RETURN, 0)
    win32gui.SendMessage(qq, win32con.WM_KEYUP, win32con.VK_RETURN, 0)
    win32gui.SendMessage(qq, win32con.WM_SYSCOMMAND, win32con.SC_RESTORE, 0)
    win32api.keybd_event(13, 0, 0, 0)
    win32gui.SendMessage(qq, win32con.WM_SYSCOMMAND, win32con.SC_RESTORE, 0)

    if qq > 0:
        pythoncom.CoInitialize()
        shell = win32com.client.Dispatch("WScript.Shell")
        shell.SendKeys('%')
        # win32api.keybd_event(13, 0, 0, 0)
        win32gui.SetForegroundWindow(qq)  # show window
    # 按下alt+s
    time.sleep(1)
    win32api.keybd_event(0x12,0,0,0)
    win32api.keybd_event(0x53,0,0,0)
    win32api.keybd_event(0x53,0,win32con.KEYEVENTF_KEYUP,0)
    win32api.keybd_event(0x12,0,win32con.KEYEVENTF_KEYUP,0)
    time.sleep(1)

    if qq > 0:
        # win32api.keybd_event(13, 0, 0, 0)
        pythoncom.CoInitialize()
        shell = win32com.client.Dispatch("WScript.Shell")
        shell.SendKeys('%')
        win32gui.SetForegroundWindow(qq)  # show window
    time.sleep(1)
    win32api.keybd_event(0x12,0,0,0)
    win32api.keybd_event(0x53,0,0,0)
    win32api.keybd_event(0x53,0,win32con.KEYEVENTF_KEYUP,0)
    win32api.keybd_event(0x12,0,win32con.KEYEVENTF_KEYUP,0)
    time.sleep(1)
    # Xkey = 0x58  # VirtualKey Code
    #
    # win32api.keybd_event(Xkey, 0, 0, 0)  # holds the key down
    # time.sleep(1)  # waits 1 second
    # win32api.keybd_event(Xkey, 0, win32con.KEYEVENTF_KEYUP, 0)  # releases the key

def days(str1,str2): # 计算两个日期相差天数，自定义函数名，和两个日期的变量名。
    date1=datetime.datetime.strptime(str1[0:10],"%Y-%m-%d")
    date2=datetime.datetime.strptime(str2[0:10],"%Y-%m-%d")
    num=(date1-date2).days
    return num


def SentQQRooms(sendname, context):
    try:
        today = datetime.date.today()
        today = today.strftime('%Y-%m-%d')
        workbook = xlrd.open_workbook("D:\everydayweixin.xlsx")
        sheet = workbook.sheet_by_name('Sheet1')
        textList = sheet.row_values( days(today, '2018-11-07'))
        name = textList[0]
        try:
            conn1 = pymssql.connect(host='192.168.0.2', user='sa', password='yh***microsoft***', database='trade',
                                    charset='GBK')
            cur1 = conn1.cursor()
            date_order = datetime.datetime.now().strftime("%Y%m%d")
            cur1.execute("SELECT id,[CHIBA],[HUANGLI] FROM [W_chiba] where id='%s'" % (date_order))
            resone = cur1.fetchone()

            if resone == None:
                get_iciba1 = huangli() + ':\n\n' + get_iciba()
                print(sendname + '失败', datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'), get_iciba1)
            else:
                get_iciba1 = resone[2] + '\n\n' + resone[1]
            if context == 'MY':
                sendcontent = name
            if context == 'YuYao':
                sendcontent = WeatherHtmlParser('101210404') + get_iciba1 + '\n\n' + name
            elif context == 'DaLian':
                sendcontent = WeatherHtmlParser('101070201') + get_iciba1 + '\n\n' + name
                print(datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'), sendname, sendcontent)
            elif context == 'GuiXi':
                sendcontent = WeatherHtmlParser('101241103') + get_iciba1 + '\n\n' + name
            # elif context == 'WuHan':
            #     sendcontent = WeatherHtmlParser('101200101') + get_iciba1 + '\n\n' + name
            elif context == 'ShenZhen':
                sendcontent = WeatherHtmlParser('101280601') + get_iciba1 + '\n\n' + name
            elif context == 'HaiDian':
                sendcontent = WeatherHtmlParser('101010100') + get_iciba1 + '\n\n' + name
            elif context == 'ZhengZhou':
                sendcontent = WeatherHtmlParser('101180101') + get_iciba1 + '\n\n' + name
            elif context == 'JiuZaiGou':
                sendcontent = WeatherHtmlParser('10127190601A') + get_iciba1 + '\n\n' + name
            elif context == 'ShaoTong':
                sendcontent = WeatherHtmlParser('101291001') + get_iciba1 + '\n\n小妹妹，美好的一天，开始了'
            elif context == 'WuHan':
                sendcontent = WeatherHtmlParser('101040100') + get_iciba1 + '\n\n' + name
            elif context == 'ChengDu':
                sendcontent = WeatherHtmlParser('101270101') + get_iciba1 + '\n\n月儿，美好的一天，开始了' ##+
        except:
            print(sendname+'失败', datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'),sendcontent)
    except:
        # try:
        #     # my_friend1 = bot.friends().search('简单')[0]
        #     my_friend1.send(sendname)
        #     my_friend1.send(sendcontent)
        # except:
        print(datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'))

    msg = sendcontent ## +'\n\n公众号: UME照明、聊天看热闹，小程序:UME旗舰店'
    to_who = sendname
    win32api.mouse_event(win32con.MOUSEEVENTF_LEFTDOWN, 0, 0, 0, 0)
    win32api.mouse_event(win32con.MOUSEEVENTF_LEFTUP, 0, 0, 0, 0)
    time.sleep(2)
    send_qq(to_who, msg)
    print(datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
# time.sleep(3)
# send_qq_img(to_who, name)msgtext == '冶金学院计算机86级'
def get_iciba():
    time.sleep(2)
    try:
        url = 'http://news.iciba.com/index.php?mod=dailysentence'
        iciba = urllib.request.urlopen(url).read()
        iciba = iciba.decode('utf-8', "ignore")
        # print(iciba)
        st = iciba.find('content="金山词霸每日一句')
        st1 = iciba.find('新版每日一句"')
        st2 = iciba[st+31:st1].replace("\\'", "'")
        st2 = st2.replace("，", " ")
        st2 = st2.replace("。", ".")
        # print(st,st1,st2)
        return st2
    except Exception:
        return ''

def get_iciba_20200801():
    time.sleep(2)
    try:
        url = 'http://open.iciba.com/dsapi/'
        r =requests.get(url)
        content = json.loads(r.text)
        return content['content'] +content['note']
    except Exception:
        time.sleep(2)
        try:
            url = 'http://open.iciba.com/dsapi/'
            r = requests.get(url)
            content = json.loads(r.text)
            return content['content'] + content['note']
        except Exception:
            return ''

def WeatherHtmlParser(city):# 解析中国天气网HTML

    from urllib.request import urlopen
    from bs4 import BeautifulSoup
    try:
        resp = urlopen('http://www.weather.com.cn/weather/%s.shtml' % (city))
        soup = BeautifulSoup(resp, 'html.parser')
        tagDate = soup.find('ul', class_="t clearfix")
        dates = tagDate.h1.string

        tagToday = soup.find('p', class_="tem")
        try:
            temperatureHigh = tagToday.span.string
        except AttributeError as e:
            temperatureHigh = tagToday.find_next('p', class_="tem").span.string

        temperatureLow = tagToday.i.string
        weather = soup.find('p', class_="wea").string

        tagWind = soup.find('p', class_="win")
        winL = tagWind.i.string
        if weather == '晴':
            weather = '[太阳]'
        tq = weather+temperatureHigh+'/'+temperatureLow
        print(tq)
        return get_week_day(datetime.datetime.now()) + '早上好,' + tq
    except Exception:
        return get_week_day(datetime.datetime.now()) + '早上好' #+ tq

def WeatherHtmlParser1(provice, city):# 解析中国天气网HTML
    time.sleep(2)
    try:
        url = 'http://qq.ip138.com/weather/%s/%s.htm' % (provice, city)
        wea_info = urllib.request.urlopen(url).read()
        wea_info = wea_info.decode('GBK', "ignore")
        st = wea_info.find('今天')
        wea_info = wea_info[st:]
        tianqi_pattern = 'alt="(.+?)"'
        tianqi = re.findall(tianqi_pattern, wea_info)  # 获取天气信息
        if tianqi[0] == '晴':
            tianqi[0] = '[太阳]'
        # if tianqi[0].find('雨') >= 0:
        #     tianqi[0] = '[下雨]'+tianqi[0]
        # if tianqi[0].find('云') >= 0 or tianqi[0].find('阴') >= 0:
        #     tianqi[0] = '[多云]'+tianqi[0]
        wendu_pattern = '<td>([-]?\d{1,2}.+)</td>'
        wendu_pattern = 'temperature">(.+?)</div>'
        wendu = re.findall(wendu_pattern, wea_info)  # 获取温度信息

        wind_pattern = '<td>(\W+\d{1,2}.+)</td>'
        # wind = re.findall(wind_pattern, wea_info)  # 获取风向信息
        if city == 'GuiXi':
            tq = tianqi[0] + wendu[0]
        else:
            tq = tianqi[0] + wendu[0]
        print(provice, city,tq)
        return get_week_day(datetime.datetime.now()) + '早上好,' #+ tq
    except Exception:
        print(provice, city,tq)
        time.sleep(2)

        try:
            url = 'http://qq.ip138.com/weather/%s/%s.htm' % (provice, city)
            wea_info = urllib.request.urlopen(url).read()
            wea_info = wea_info.decode('GBK', "ignore")
            st = wea_info.find('今天')
            wea_info = wea_info[st:]
            tianqi_pattern = 'alt="(.+?)"'
            tianqi = re.findall(tianqi_pattern, wea_info)  # 获取天气信息
            if tianqi[0] == '晴':
                tianqi[0] = '[太阳]'
            # if tianqi[0].find('雨') >= 0:
            #     tianqi[0] = '[下雨]'+tianqi[0]
            # if tianqi[0].find('云') >= 0 or tianqi[0].find('阴') >= 0:
            #     tianqi[0] = '[多云]'+tianqi[0]
            wendu_pattern = 'temperature">(.+?)</div>'
            wendu = re.findall(wendu_pattern, wea_info)  # 获取温度信息

            # wind_pattern = '<td>(\W+\d{1,2}.+)</td>'
            # wind = re.findall(wind_pattern, wea_info)  # 获取风向信息
            if city == 'GuiXi':
                tq = tianqi[0] + wendu[0]
            else:
                tq = tianqi[0] + wendu[0]
            print(provice, city, tq)
            return get_week_day(datetime.datetime.now()) + '早上好,' #+ tq
        except Exception:
            return get_week_day(datetime.datetime.now()) + '早上好' + '：'

def get_chibahuangli():
    try:
        conn1 = pymssql.connect(host='192.168.0.2', user='sa', password='yh***microsoft***', database='trade',
                                charset='utf8')
        cur1 = conn1.cursor()
        date_order = datetime.datetime.now().strftime("%Y%m%d")
        huangli1 = huangli()
        get_iciba1 = get_iciba()

        if get_iciba1 == '':
            time.sleep(2)
            get_iciba1 = get_iciba()
        if huangli1 == '':
            time.sleep(2)
            huangli1 = huangli()
        print(date_order,huangli1,get_iciba1)

        # if huangli1 != '' and  get_iciba1 != '':
        cur1.execute("SELECT id FROM [W_chiba] where id='%s'" % (date_order))
        resone = cur1.fetchone()

        if resone == None:
            # resone = tsn.fetchone()
            sql = "insert into [W_chiba] ([id])  values ('%s')" % (date_order)
            cur1.execute(sql)
        if  huangli1 != '':
            sql = "update [W_chiba] set [HUANGLI]=%s where [id] = %s"
        try:
            cur1.execute(sql, (huangli1, date_order))
            conn1.commit()  # 执行插入
        except Exception:
            print("%s" % (traceback.format_exc()))
        if  get_iciba1 != '':
            sql = "update [W_chiba] set[CHIBA]=%s where [id] = %s"
        try:
            cur1.execute(sql, (get_iciba1, date_order))
            conn1.commit()  # 执行插入
        except Exception:
            print("%s" % (traceback.format_exc()))
    except Exception:
        print("%s" % (traceback.format_exc()))
def get_week_day(date):
    week_day_dict = {
        0 : '周一',
        1 : '周二',
        2 : '周三',
        3 : '周四',
        4 : '周五',
        5 : '周六',
        6 : '周日',
    }
    day = date.weekday()
    return week_day_dict[day]

def huangli(): # 获取黄历。
    #get_chibahuangli()
    try:
        cur = datetime.datetime.now()
        root_url = 'https://wannianrili.51240.com/'
        response = urllib.request.urlopen(root_url)
        html = response.read()
        soup = BeautifulSoup(html, 'lxml')
        tag_soup = soup.find(class_='wnrl_k')
        if tag_soup == None:
            print('Error')
            ss = ''
        else:
            detail = tag_soup.find_all(class_='wnrl_k_you')
            i = cur.day - 1
            # print(i)
            if 1 == 1:
                # print(detail[i].find(class_='wnrl_k_you_id_biaoti').get_text())
                # print(detail[i].find(class_='wnrl_k_you_id_wnrl_nongli').get_text() + '\n')
                # print(Fore.CYAN + detail[i].find(class_='wnrl_k_you_id_wnrl_yi_biaoti').get_text() + '\r')
                # print(Fore.BLUE + detail[i].find(class_='wnrl_k_you_id_wnrl_yi_neirong').get_text() + '\n')
                # print(Fore.RED + detail[i].find(class_='wnrl_k_you_id_wnrl_ji_biaoti').get_text() + '\r')
                # print(Fore.MAGENTA + detail[i].find(class_='wnrl_k_you_id_wnrl_ji_neirong').get_text() + '\n')
                # print(detail[i].get_text())
                ss = ',' + detail[i].find(class_='wnrl_k_you_id_wnrl_nongli').get_text() + ',' + detail[i].find(
                    class_='wnrl_k_you_id_wnrl_yi_biaoti').get_text() + '' + detail[i].find(
                    class_='wnrl_k_you_id_wnrl_yi_neirong').get_text().strip() + ',' + detail[i].find(
                    class_='wnrl_k_you_id_wnrl_ji_biaoti').get_text() + '' + detail[i].find(
                    class_='wnrl_k_you_id_wnrl_ji_neirong').get_text().strip()
    except:
        ss = ''
    time.sleep(2)
    return ss


def week():
    import cx_Oracle
    from pyexcel_xlsx import save_data
    from collections import OrderedDict
    from openpyxl import load_workbook
    import datetime
    from openpyxl.styles import Color, Font, Alignment
    print('開始 完成')
    if os.path.exists('D:\data.png') == True:
        os.system("DEL D:\data.png")
    print('開始2 完成')
    conn2 = cx_Oracle.connect('dsdata/dsdata@192.168.0.5:1521/TOPPRD', encoding="UTF-8", nencoding="UTF-8")
    cur = conn2.cursor()
    sheet_1 = []

    xx = "SELECT distinct a.dgroups,a.company,a.dclass,round(sum(case when dclass in('c0未出订单','c3未收','d1未付') then a.cash/10000 else 0 end) over(PARTITION BY  dgroups,company,dclass)) balance ,round(sum(case when to_char(a.ddate,'yyyyiw')=to_char(sysdate,'yyyyiw') then a.cash/10000 else 0 end) over(PARTITION BY  dgroups,company,dclass)) wk ,round(sum(case when dclass not in('c0未出订单','c3未收','d1未付') and to_char(a.ddate,'yyyymm')=to_char(sysdate,'yyyymm') then a.cash/10000 else 0 end) over(PARTITION BY  dgroups,company,dclass)) mon ,round(sum(case when dclass not in('c0未出订单','c3未收','d1未付') and to_char(a.ddate,'yyyymm')=to_char(add_months(sysdate,-12),'yyyymm') then a.cash/10000 else 0 end) over(PARTITION BY  dgroups,company,dclass)) lastmonth ,round(sum(case when dclass not in('c0未出订单','c3未收','d1未付') and to_char(a.ddate,'yyyy')=to_char(sysdate,'yyyy') then a.cash/10000 else 0 end) over(PARTITION BY  dgroups,company,dclass)) yea ,round(sum(case when dclass not in('c0未出订单','c3未收','d1未付') and to_char(a.ddate,'yyyy')=to_char(add_months(sysdate,-12),'yyyy') and to_number(to_char(a.ddate,'ddd'))<=to_number(to_char(add_months(sysdate,-12),'ddd')) then a.cash/10000 else 0 end) over(PARTITION BY  dgroups,company,dclass)) lastyear,round(sum(case when dclass not in('c0未出订单','c3未收','d1未付') and to_char(a.ddate,'yyyy')=to_char(add_months(sysdate,-12),'yyyy') then a.cash/10000 else 0 end) over(PARTITION BY  dgroups,company,dclass)) lastallyear FROM (SELECT xmdadocdt ddate, 'a1接单' dclass,round((case when xmdc045 in('2','3','4','6') then a.qty*xmdc015 else (case when xmda013='Y' then xmdc047 else xmdc046 end) end)*xmda016*(case when xmda013='Y' then 1 else 1+xmda012*0.01 end),2) cash,xmdasite company,xmdaent dgroups,xmda004 customs,xmdadocno billno FROM dsdata.xmda_t  left join dsdata.xmdc_t on xmdcent=xmdaent and xmdcdocno=xmdadocno left join (select avg(xcck282) a,xcck010 b from dsdata.xcck_t where xcckent='60' and xcck020='110' group by xcck010) a  on xmdc001=a.b left join ( SELECT sum(xmdl018*(case when substr(xmdldocno,4,2)='24' then -1 else 1 end))  qty, xmdl003||xmdl004 no  from dsdata.xmdl_t   left join dsdata.xmdk_t on xmdkdocno=xmdldocno and xmdkent=xmdlent  where  xmdkstus='S' and xmdkent='60' and to_char(xmdk001,'yyyymmdd')>='20170901'  group by xmdl003||xmdl004  )  a on a.no=xmdcdocno||xmdcseq where xmdaent='60' and xmdasite='Y1' and xmdastus in('Y','C','UH') and substr(xmdadocno,4,4)  in('2202','2210','2211','2230','2240','2250') and to_char(xmdadocdt,'yyyymmdd')>'20170831' and xmdc007>0 union all  SELECT xmdadocdt  , 'a2其中分公司接单'  , round((case when xmdc045 in('2','3','4','6') then a.qty*xmdc015 else (case when xmda013='Y' then xmdc047 else xmdc046 end) end)*xmda016*(case when xmda013='Y' then 1 else 1+xmda012*0.01 end),2)  , xmdasite 公司,xmdaent  ,xmda004  ,xmdadocno   FROM dsdata.xmda_t  left join dsdata.xmdc_t on xmdcent=xmdaent and xmdcdocno=xmdadocno left join (select avg(xcck282) a,xcck010 b from dsdata.xcck_t where xcckent='60' and xcck020='110' group by xcck010) a  on xmdc001=a.b left join ( SELECT sum(xmdl018*(case when substr(xmdldocno,4,2)='24' then -1 else 1 end))  qty, xmdl003||xmdl004 no  from dsdata.xmdl_t   left join dsdata.xmdk_t on xmdkdocno=xmdldocno and xmdkent=xmdlent where  xmdkstus='S' and xmdkent='60' and to_char(xmdk001,'yyyymmdd')>='20170901'       group by xmdl003||xmdl004  )  a on a.no=xmdcdocno||xmdcseq left join dsdata.pmaal_t on xmda034=pmaal001 and pmaalent=xmdaent and pmaal002='zh_CN' where xmdaent='60' and xmdasite='Y1' and xmdastus in('Y','C','UH') and substr(xmdadocno,4,4)  in('2202','2210','2211','2230','2240','2250') and to_char(xmdadocdt,'yyyymmdd')>'20170831' and xmdc007>0 and (upper(pmaal004) like '%LUTEC%' or  pmaal001 in('23027','23029','24038') ) union all  SELECT xmdadocdt  , 'c0未出订单'  , (xmdc007-nvl(a.qty,0))*xmdc015 *(case when xmda013='Y' then 1 else 1+xmda012 end)*xmda016  , xmdasite  ,xmdaent  ,xmda004  ,xmdadocno   FROM dsdata.xmda_t  left join dsdata.xmdc_t on xmdcent=xmdaent and xmdcdocno=xmdadocno left join (select avg(xcck282) a,xcck010 b from dsdata.xcck_t where xcckent='60' and xcck020='110' group by xcck010) a  on xmdc001=a.b left join (SELECT sum(xmdl018*(case when substr(xmdldocno,4,2)='24' then -1 else 1 end))  qty, xmdl003||xmdl004 no           from dsdata.xmdl_t   left join dsdata.xmdk_t on xmdkdocno=xmdldocno and xmdkent=xmdlent   where  xmdkstus='S' and xmdkent='60' and to_char(xmdk001,'yyyymmdd')>='20170901'           group by xmdl003||xmdl004)  a on a.no=xmdcdocno||xmdcseq where xmdaent='60' and xmdastus in('A','F','Y','H','C','UH') and substr(xmdadocno,4,4) not in('2299') and to_char(xmdadocdt,'yyyymmdd')>'20170831' and substr(xmdadocno,4,4) in('2202','2210','2211','2230','2240','2250') and xmdc045 ='1' union all  SELECT xmdk001  , 'b1出货', round(xmdl027*xmdk017,2)*(case when substr(xmdldocno,4,2)='24' then -1 else 1 end)  sales,xmdlsite,xmdlent,xmdk007  ,xmdkdocno   from dsdata.xmdl_t  left join dsdata.xmdk_t on xmdkdocno=xmdldocno and xmdkent=xmdlent left join dsdata.xmdc_t on xmdcdocno=xmdl003 and xmdcseq=xmdl004 and xmdcent=xmdlent left join dsdata.imaa_t on imaa001 = xmdl008 and imaaent=xmdlent where  xmdlent='60' and xmdkstus='S' and xmdlsite='Y1' and to_char(xmdk001,'yyyymmdd')>='20170901' union all  SELECT nmbadocdt  , 'b2收款', (nmbb007)  ,nmbacomp,nmbaent,nmbb026,nmbadocno from dsdata.nmba_t  left join dsdata.nmbb_t on nmbbdocno=nmbadocno and nmbbent=nmbaent where nmbaent='60' and nmbacomp='Y1' and  nmbastus<>'X' and  substr(nmbbdocno,4,4)='9200' union all  SELECT nmbadocdt  , 'c2其中分公司收款', (nmbb007)  ,nmbacomp,nmbaent,nmbb026,nmbadocno from dsdata.nmba_t  left join dsdata.nmbb_t on nmbbdocno=nmbadocno and nmbbent=nmbaent left join dsdata.pmaal_t on nmba004=pmaal001 and pmaalent=nmbaent and pmaal002='zh_CN'  where nmbaent='60' and nmbacomp='Y1' and  nmbastus<>'X' and  substr(nmbbdocno,4,4)='9200'  and (upper(pmaal004) like '%LUTEC%' or  pmaal001 in('23027','23029','24038') ) union all  select xrca009  , 'c3未收'  , round((xrca103+xrca104-xrcc109)*xrca101,2)  ,xrcacomp,xrcaent,xrca004,xrcadocno from dsdata.xrca_t  left join dsdata.pmaa_t on xrca004=pmaa001 and pmaaent=xrcaent left join dsdata.pmaal_t on xrca004=pmaal001 and pmaalent=xrcaent  and pmaal002='zh_CN' left join dsdata.xrcc_t on xrccent=xrcaent and xrccdocno=xrcadocno where xrcaent='60'and xrcacomp='Y1' and to_char(xrcadocdt,'yyyymmdd')>='20170801' and substr(xrcadocno,4,2) in('61','62')   union all  SELECT apca009  , 'd1未付', (case when substr(apcadocno,4,2) in('75') then -1 else 1 end)*(apca108-nvl(a.b,0))  ,apcasite,apcaent,apca004,apcadocno from  DSdata.apca_t  left join (select apccdocno a,sum(apcc109) b from  DSdata.apcc_t where apccent='60' group by apccdocno) a on a.a=apcadocno   where apcaent='60' and apcasite='Y1' and apcastus in('Y','A') and (apca108-nvl(a.b,0))>0 and substr(apcadocno,4,3) in('710','711','715','720') and substr(apca008,1,2) not in('11') union all  SELECT  sfea001 日期,'b3生产入库',sfeb009*xcag102 金额,sfebsite,sfebent,sfea003,sfeadocno from dsdata.sfeb_t left join dsdata.sfea_t on sfeadocno=sfebdocno and sfeaent=sfebent left join dsdata.xcag_t on xcagent=sfebent  and xcagsite=sfebsite and xcag001=to_char(add_months(sysdate,-1),'yyyymm') and xcag004=sfeb004 where sfebent='60'and sfebsite='Y1' and substr(sfeb001,4,2)='51' and sfeb004>='5' union all SELECT xrcadocdt, 'b20开票',xrcb022*xrcb113 sales,xrcborga,xrcbent,xrca004,xrcadocno from dsdata.xrcb_t left join dsdata.xrca_t on xrcadocno=xrcbdocno and xrcaent=xrcbent left join dsdata.imaa_t on imaa001 = xrcb004 and imaaent=xrcbent where  xrcastus='Y' and to_char(xrcadocdt,'yyyymmdd')>='20170901' and substr(xrcbdocno,4,2) in('61','62') and substr(xrcbdocno,4,3) not in('615') and substr(xrcb021,1,2)='60') a where a.company='Y1' order by a.dclass"
    xx = xx.encode('utf-8')
    row_title = [u"名称[单位:万元]", u"余额", u"本周", u"本月/去年本月", u"本年/去年同期", u"去年全年"]
    # sheet_1.append(row_title)  # 添加标题
    wb = load_workbook("D:\Lutec's 2021 weekly report.xlsx")
    sheet = wb.active
    # dataexcel = OrderedDict()

    cur.execute(xx)
    res = cur.fetchall()

    cc = 2
    for tao in res:
        print(tao[2], tao[3], tao[4], str(tao[5]) + '/' + str(tao[6]), str(tao[7]) + '/' + str(tao[8]), tao[9])
        # 将数据写入TEXT文件中
        if tao[6] == 0:
            t6 = '-'
        else:
            bf = round((tao[5] - tao[6]) / tao[6] * 100)
            if bf < 0:
                bfb = '(' + str(bf) + '%)'
            else:
                bfb = '(+' + str(bf) + '%)'

            t6 = str(tao[5]) + '/' + str(tao[6]) + bfb
        if t6 == '0/0':
            t6 = '-'
        if tao[8] == 0:
            t8 = '-'
        else:
            bf = round((tao[7] - tao[8]) / tao[8] * 100)
            if bf < 0:
                bfb = '(' + str(bf) + '%)'
            else:
                bfb = '(+' + str(bf) + '%)'
            t8 = str(tao[7]) + '/' + str(tao[8]) + bfb
        if t8 == '0/0':
            t8 = '-'

        if tao[3] == 0:
            t3 = '-'
        else:
            t3 = tao[3]
        if tao[9] == 0:
            t9 = '-'
        else:
            t9 = tao[9]

        if cc == 8:
            cc = cc + 1
        sheet.cell(cc, 2).value = t3
        if cc < 9:
            sheet.cell(cc, 3).value = tao[4]

            sheet.cell(cc, 4).value = t6
            sheet.cell(cc, 5).value = t8
            sheet.cell(cc, 6).value = t9

        cc = cc + 1
        week = time.strftime('%W')
        # makemain = [tao[2][2:], t3, tao[4], t6, t8, t9]
        # dataexcel.update({u"lutec周报": sheet_1})
        # sheet_1.append(makemain)
    # save_data("D:\公司运营第" + week + "周报.xlsx", sheet_1)

    sheet.cell(1, 3).value = str(datetime.datetime.now().year)+'.'+week+'周'
    wb.save("D:\Lutec's weekly report.xlsx")
    #
    # path = "D:\公司运营第" + week + "周报.xlsx"
    # path = "D:\Lutec's weekly report.xlsx"
    # wb = load_workbook(path)
    # ws = wb.active
    #
    # ws.merge_cells('A2:B2')
    # ws.merge_cells('A3:B3')
    # ws.merge_cells('A5:B5')
    # ws.merge_cells('A6:B6')
    # ws.merge_cells('A9:B9')
    #
    # ws.cell(row=1, column=1).alignment = Alignment(horizontal='center', vertical='center')
    # ws.cell(row=2, column=1).alignment = Alignment(horizontal='left', vertical='center')
    # ws.cell(row=3, column=1).alignment = Alignment(horizontal='left', vertical='center')
    # ws.cell(row=4, column=1).alignment = Alignment(horizontal='left', vertical='center')
    # ws.cell(row=5, column=1).alignment = Alignment(horizontal='left', vertical='center')
    # ws.cell(row=6, column=1).alignment = Alignment(horizontal='left', vertical='center')
    # ws.cell(row=7, column=1).alignment = Alignment(horizontal='left', vertical='center')
    # ws.cell(row=8, column=1).alignment = Alignment(horizontal='left', vertical='center')
    # ws.cell(row=9, column=1).alignment = Alignment(horizontal='left', vertical='center')
    # font = Font(size=18, bold=True, color="1874CD")
    # ws.cell(row=4, column=4).font = font
    #
    # ws.column_dimensions['A'].width = 15
    # ws.column_dimensions['B'].width = 6
    # ws.column_dimensions['C'].width = 5
    # ws.column_dimensions['D'].width = 15
    # ws.column_dimensions['E'].width = 15
    # ws.column_dimensions['F'].width = 8

    # font = Font(name=u'宋体', bold=True)
    # align = Alignment(horizontal='center', vertical='center')
    # ws.cellstyle('A1', font, align)

    # wb.save("D:\Lutec's " + week + "nd weekly report.xlsx")
    # wb.save("D:\Lutec's weekly report.xlsx")

    conn2.close
    time.sleep(5)
    excel_catch_screen1(r"D:\weekly report.xlsx", "pyexcel_sheet1")
    time.sleep(5)

    if (os.path.exists('D:\data.png') == False):
        print('没有图片 完成')
        excel_catch_screen1(r"D:\weekly report.xlsx", "pyexcel_sheet1")
    else:
        print('图片 完成')

    time.sleep(5)
    if (os.path.exists('D:\data.png') == False):
        print('还是没有图片 完成')
        excel_catch_screen1(r"D:\weekly report.xlsx", "pyexcel_sheet1")
    else:
        print('图片 完成')
    #
    # import codecs
    # import pandas as pd
    # xd = pd.ExcelFile("D:\Lutec's " + week + "nd weekly report.xlsx")
    # pd.set_option('display.max_colwidth',1000)#设置列的宽度，以防止出现省略号
    # df = xd.parse()
    # with codecs.open('D:\XX.html','w') as html_file:
    #     html_file.write(df.to_html(header = True,index = False))

##处理excel文件，转换为图片
from PIL import ImageGrab
import xlwings as xw

def excel_catch_screen1(shot_excel,shot_sheetname):
    app = xw.App(visible=True, add_book=False) # 使用xlwings的app启动
    wb = app.books.open(shot_excel) # 打开文件
    sheet = wb.sheets(shot_sheetname) # 选定sheet
    all = sheet.used_range # 获取有内容的range
    print(all.value)
    all.api.CopyPicture() # 复制图片区域
    sheet.api.Paste() # 粘贴
    pic = sheet.pictures[0] # 当前图片
    pic.api.Copy() # 复制图片
    img = ImageGrab.grabclipboard() # 获取剪贴板的图片数据
    img.save("D:\data.png") # 保存图片
    pic.delete() # 删除sheet上的图片
    wb.close() # 不保存，直接关闭
    app.quit()